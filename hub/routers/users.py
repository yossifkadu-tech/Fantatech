"""
users.py — User management router
Primary database: users.csv (simple, always works)
Secondary export:  users.xlsx  (for Excel reporting)

Endpoints:
  GET  /api/users/types            → list all user types with permissions
  GET  /api/users/list             → list all users (no PIN)
  POST /api/users/add              → add a new user to CSV
  PUT  /api/users/{id}             → update a user
  DELETE /api/users/{id}           → delete a user (soft-delete: status=deleted)
  POST /api/users/register         → legacy: register + append to CSV
  POST /api/users/login            → look up username, return record (no PIN)
  GET  /api/users/export           → download users.xlsx
  GET  /api/users/count            → number of active users
"""
import os
import csv
import json
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List

router = APIRouter()

# ── File paths ───────────────────────────────────────────────────────────────
BASE      = os.path.dirname(os.path.dirname(__file__))
CSV_PATH  = os.path.join(BASE, "users.csv")
TYPES_PATH= os.path.join(BASE, "user_types.csv")
XLSX_PATH = os.path.join(BASE, "users.xlsx")

# ── CSV columns ──────────────────────────────────────────────────────────────
COLUMNS = ["id", "type", "name", "username", "email",
           "address", "plan", "pin", "status", "created_at", "notes"]

# ── Pydantic models ───────────────────────────────────────────────────────────
class UserAdd(BaseModel):
    type:     str
    name:     str
    username: str
    email:    Optional[str] = ""
    address:  Optional[str] = ""
    plan:     Optional[str] = "free"
    pin:      Optional[str] = ""
    notes:    Optional[str] = ""

class UserUpdate(BaseModel):
    type:     Optional[str] = None
    name:     Optional[str] = None
    email:    Optional[str] = None
    address:  Optional[str] = None
    plan:     Optional[str] = None
    pin:      Optional[str] = None
    status:   Optional[str] = None
    notes:    Optional[str] = None

class UserRegister(BaseModel):
    plan:        str
    name:        str
    username:    str
    email:       str
    address:     Optional[str] = ""
    card_holder: Optional[str] = ""
    card_number: Optional[str] = ""
    card_expiry: Optional[str] = ""
    card_cvv:    Optional[str] = ""

class UserLogin(BaseModel):
    username: str


# ── CSV helpers ───────────────────────────────────────────────────────────────

def _ensure_csv():
    """Create users.csv with header if it doesn't exist."""
    if not os.path.exists(CSV_PATH):
        with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=COLUMNS)
            w.writeheader()

def _read_users() -> List[dict]:
    _ensure_csv()
    with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def _write_users(rows: List[dict]):
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)

def _next_id(rows: List[dict]) -> int:
    if not rows:
        return 1
    return max((int(r.get("id", 0) or 0) for r in rows), default=0) + 1

def _safe(row: dict) -> dict:
    """Return user dict without PIN."""
    return {k: v for k, v in row.items() if k != "pin"}

def _read_types() -> List[dict]:
    if not os.path.exists(TYPES_PATH):
        return []
    with open(TYPES_PATH, "r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/types")
async def get_user_types():
    """Return all defined user types with their permissions."""
    return {"types": _read_types()}


@router.get("/list")
async def list_users():
    """List all active users (PIN excluded)."""
    rows = _read_users()
    active = [_safe(r) for r in rows if r.get("status", "active") != "deleted"]
    return {"users": active, "total": len(active)}


@router.post("/add")
async def add_user(data: UserAdd):
    """Add a new user to users.csv."""
    rows = _read_users()
    # Username must be unique
    if any(r.get("username","").strip().lower() == data.username.strip().lower()
           for r in rows if r.get("status") != "deleted"):
        raise HTTPException(status_code=409, detail="Username already exists")

    new_row = {
        "id":         _next_id(rows),
        "type":       data.type,
        "name":       data.name,
        "username":   data.username.strip(),
        "email":      data.email or "",
        "address":    data.address or "",
        "plan":       data.plan or "free",
        "pin":        data.pin or "",
        "status":     "active",
        "created_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "notes":      data.notes or "",
    }
    rows.append(new_row)
    _write_users(rows)
    _rebuild_xlsx(rows)
    return {"ok": True, "user": _safe(new_row)}


@router.put("/{user_id}")
async def update_user(user_id: int, data: UserUpdate):
    """Update fields on an existing user."""
    rows = _read_users()
    for row in rows:
        if str(row.get("id")) == str(user_id):
            if data.type    is not None: row["type"]    = data.type
            if data.name    is not None: row["name"]    = data.name
            if data.email   is not None: row["email"]   = data.email
            if data.address is not None: row["address"] = data.address
            if data.plan    is not None: row["plan"]    = data.plan
            if data.pin     is not None: row["pin"]     = data.pin
            if data.status  is not None: row["status"]  = data.status
            if data.notes   is not None: row["notes"]   = data.notes
            _write_users(rows)
            _rebuild_xlsx(rows)
            return {"ok": True, "user": _safe(row)}
    raise HTTPException(status_code=404, detail="User not found")


@router.delete("/{user_id}")
async def delete_user(user_id: int):
    """Soft-delete a user (status = deleted)."""
    rows = _read_users()
    for row in rows:
        if str(row.get("id")) == str(user_id):
            row["status"] = "deleted"
            _write_users(rows)
            return {"ok": True}
    raise HTTPException(status_code=404, detail="User not found")


@router.post("/login")
async def login_user(data: UserLogin):
    """
    Look up a username in users.csv and return their record (no PIN).
    Used by the app after reinstall to restore a session.
    """
    rows = _read_users()
    for row in rows:
        if (row.get("status", "active") != "deleted" and
                str(row.get("username", "")).strip().lower()
                == data.username.strip().lower()):
            return {"ok": True, "user": _safe(row)}
    raise HTTPException(status_code=404, detail="Username not found")


@router.post("/register")
async def register_user(data: UserRegister):
    """
    Legacy registration endpoint (called from app sign-up form).
    Adds user to CSV and rebuilds xlsx.
    """
    rows = _read_users()
    # Don't duplicate if username already exists
    existing = next(
        (r for r in rows
         if r.get("username","").strip().lower() == data.username.strip().lower()
         and r.get("status","active") != "deleted"),
        None
    )
    if existing:
        return {"ok": True, "message": "User already registered", "user": _safe(existing)}

    raw_card = (data.card_number or "").replace(" ", "").replace("-", "")
    last4    = raw_card[-4:] if len(raw_card) >= 4 else ""
    notes    = f"plan={data.plan}"
    if last4:
        notes += f" | card_last4={last4} | expiry={data.card_expiry}"

    new_row = {
        "id":         _next_id(rows),
        "type":       "owner",          # default type for app registrations
        "name":       data.name,
        "username":   data.username.strip(),
        "email":      data.email,
        "address":    data.address or "",
        "plan":       data.plan,
        "pin":        "",
        "status":     "active",
        "created_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "notes":      notes,
    }
    rows.append(new_row)
    _write_users(rows)
    _rebuild_xlsx(rows)
    return {"ok": True, "message": "Registered successfully"}


@router.get("/export")
async def export_users():
    """Download users.xlsx."""
    _rebuild_xlsx(_read_users())
    if not os.path.exists(XLSX_PATH):
        raise HTTPException(status_code=500, detail="Could not generate xlsx")
    return FileResponse(
        path=XLSX_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="fantatech-users.xlsx",
    )


@router.get("/export-csv")
async def export_csv():
    """Download users.csv directly."""
    _ensure_csv()
    return FileResponse(
        path=CSV_PATH,
        media_type="text/csv",
        filename="fantatech-users.csv",
    )


@router.get("/count")
async def users_count():
    rows = _read_users()
    return {"count": sum(1 for r in rows if r.get("status", "active") != "deleted")}


# ── xlsx rebuild (optional, requires openpyxl) ───────────────────────────────

def _rebuild_xlsx(rows: List[dict]):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "משתמשים"

        # Header
        display_cols = [c for c in COLUMNS if c != "pin"]
        header_fill = PatternFill("solid", fgColor="0F172A")
        header_font = Font(bold=True, color="38BDF8", size=11)
        for ci, col in enumerate(display_cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for ri, row in enumerate(rows, start=2):
            fill = PatternFill("solid", fgColor="1E293B" if ri % 2 == 0 else "0F172A")
            font = Font(color="F1F5F9", size=10)
            for ci, col in enumerate(display_cols, start=1):
                cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="right")

        # Column widths
        widths = [6, 14, 20, 18, 28, 24, 10, 10, 12, 18, 30]
        for i, w in enumerate(widths[:len(display_cols)], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        wb.save(XLSX_PATH)
    except Exception as e:
        print(f"[users] xlsx rebuild error: {e}")
